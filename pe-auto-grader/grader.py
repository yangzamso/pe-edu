"""
엑셀 자동채점 시스템 - 채점 로직 모듈 (안전한 시트 생성 방식 적용)
"""
import openpyxl
from typing import Optional
import io
import pandas as pd
from openpyxl.utils import get_column_letter
from copy import copy

class ExcelGrader:
    """엑셀 파일을 분석하고 채점하는 클래스"""
    
    # 색상 이름 매핑 (RGB 값)
    # 초록: B6D7A8 (182, 215, 168)
    # 노랑: FFE599 (255, 229, 153)
    # 빨강: EA9999 (234, 153, 153)
    
    # 객관식 배점 (E~N열)
    OBJECTIVE_SCORES = {'초록': 2.5, '노랑': 1.25, '빨강': 0}
    # 주관식 배점 (P~AD열)
    SUBJECTIVE_SCORES = {'초록': 5.0, '노랑': 2.5, '빨강': 0}
    
    # 열 범위 (0-based index)
    OBJECTIVE_COLS = list(range(4, 14))  # E~N열 (4~13)
    OBJECTIVE_SUM_COL = 14               # O열 (14)
    SUBJECTIVE_COLS = list(range(15, 30))  # P~AD열 (15~29)
    SUBJECTIVE_SUM_COL = 30              # AE열 (30)
    TOTAL_SUM_COL = 31                   # AF열 (31)
    
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.workbook = None
        self.answer_sheet = None
        
    def load_workbook(self) -> bool:
        """엑셀 파일 로드"""
        try:
            self.workbook = openpyxl.load_workbook(self.file_path, data_only=False)
            self.answer_sheet = self.workbook[self.workbook.sheetnames[0]]
            return True
        except Exception as e:
            print(f"파일 로드 실패: {e}")
            return False
    
    def get_cell_color(self, cell) -> Optional[str]:
        """셀의 배경색을 추출하여 색상 이름으로 반환"""
        fill = cell.fill
        if not fill or not fill.start_color:
            return None
            
        if fill.start_color.type == 'rgb':
            rgb = fill.start_color.rgb
            if rgb and len(rgb) >= 6:
                if len(rgb) == 8:
                    rgb = rgb[2:]
                try:
                    r = int(rgb[0:2], 16)
                    g = int(rgb[2:4], 16)
                    b = int(rgb[4:6], 16)
                    return self._identify_color(r, g, b)
                except ValueError:
                    return None
        return None
    
    def _identify_color(self, r: int, g: int, b: int) -> Optional[str]:
        """RGB 값으로 색상 식별 (허용 오차 ±30)"""
        def color_distance(r1, g1, b1, r2, g2, b2):
            return ((r1-r2)**2 + (g1-g2)**2 + (b1-b2)**2) ** 0.5
        
        if color_distance(r, g, b, 182, 215, 168) < 30: return '초록'
        if color_distance(r, g, b, 255, 229, 153) < 30: return '노랑'
        if color_distance(r, g, b, 234, 153, 153) < 30: return '빨강'
        
        # 일반적인 색상군 백업
        if g > 200 and r < 180: return '초록'
        if r > 200 and g > 200: return '노랑'
        if r > 200 and g < 180: return '빨강'
        return None

    def analyze_answer_sheet(self) -> pd.DataFrame:
        """UI 표시용 데이터프레임 생성"""
        if not self.answer_sheet:
            return pd.DataFrame()

        results = []
        for row in self.answer_sheet.iter_rows(min_row=2, max_row=self.answer_sheet.max_row):
            if len(row) <= 2: continue
            
            student_name = row[2].value
            if not student_name or str(student_name).strip() == '':
                continue
                
            # [1] 객관식 채점
            obj_score_sum = 0
            for col_idx in self.OBJECTIVE_COLS:
                if col_idx < len(row):
                    cell = row[col_idx]
                    color = self.get_cell_color(cell)
                    score = self.OBJECTIVE_SCORES.get(color, 0)
                    obj_score_sum += score

            # [2] 주관식 채점
            subj_score_sum = 0
            for col_idx in self.SUBJECTIVE_COLS:
                if col_idx < len(row):
                    cell = row[col_idx]
                    color = self.get_cell_color(cell)
                    score = self.SUBJECTIVE_SCORES.get(color, 0)
                    subj_score_sum += score
            
            results.append({
                '행번호': row[0].row,
                '학생명': student_name,
                '객관식(25점)': obj_score_sum,
                '주관식(75점)': subj_score_sum,
                '총점(100점)': obj_score_sum + subj_score_sum
            })
            
        df = pd.DataFrame(results)
        if not df.empty:
            df = df.sort_values('행번호').drop(columns=['행번호']).reset_index(drop=True)
        return df

    def generate_scored_excel(self) -> io.BytesIO:
        """
        채점 결과 파일 생성 (안전한 방식)
        copy_worksheet 대신 create_sheet 사용
        """
        output = io.BytesIO()
        
        # 워크북 안전 로드
        temp_buffer = io.BytesIO()
        self.workbook.save(temp_buffer)
        temp_buffer.seek(0)
        
        new_wb = openpyxl.load_workbook(temp_buffer)
        
        # [중요] 엑셀 파일 손상(table1.xml 오류) 방지
        # 원본 파일에 있는 '표(Table)' 정의가 openpyxl 저장 시 충돌을 일으키므로 강제 제거
        for ws in new_wb.worksheets:
            if hasattr(ws, "tables"):
                ws.tables.clear()
            if hasattr(ws, "_tables"):
                ws._tables.clear()
        
        source_sheet = new_wb[self.answer_sheet.title]
        
        # '채점결과' 시트 생성 (중복 시 번호 부여)
        base_name = "채점결과"
        target_sheet_name = base_name
        counter = 1
        
        while target_sheet_name in new_wb.sheetnames:
            counter += 1
            target_sheet_name = f"{base_name}({counter})"
            
        target_sheet = new_wb.create_sheet(target_sheet_name)
        
        # 헤더 텍스트 입력 (O, AE, AF)
        for sheet in [source_sheet, target_sheet]:
             sheet.cell(row=1, column=self.OBJECTIVE_SUM_COL + 1).value = "객관식"
             sheet.cell(row=1, column=self.SUBJECTIVE_SUM_COL + 1).value = "주관식"
             sheet.cell(row=1, column=self.TOTAL_SUM_COL + 1).value = "총합"

        # 1. 원본 데이터 전체 복사 (값만 복사하여 객체 충돌 방지)
        max_r = source_sheet.max_row
        max_c = source_sheet.max_column
        
        for r in range(1, max_r + 1):
            for c in range(1, max_c + 1):
                source_cell = source_sheet.cell(row=r, column=c)
                target_cell = target_sheet.cell(row=r, column=c)
                target_cell.value = source_cell.value
                # 필요시 스타일 복사 로직 추가 가능하지만, 
                # 오류 방지를 위해 일단 값만 복사

        # 2. 채점 및 수식 적용
        for row_idx in range(2, max_r + 1):
            student_name = source_sheet.cell(row=row_idx, column=3).value
            if not student_name or str(student_name).strip() == '':
                continue

            # [점수 기입]
            # 객관식
            for col_idx in self.OBJECTIVE_COLS:
                source_cell = source_sheet.cell(row=row_idx, column=col_idx + 1)
                target_cell = target_sheet.cell(row=row_idx, column=col_idx + 1)
                
                color = self.get_cell_color(source_cell)
                score = self.OBJECTIVE_SCORES.get(color, 0)
                target_cell.value = score
                target_cell.fill = copy(source_cell.fill)  # [추가] 배경색 복사

            # 주관식
            for col_idx in self.SUBJECTIVE_COLS:
                source_cell = source_sheet.cell(row=row_idx, column=col_idx + 1)
                target_cell = target_sheet.cell(row=row_idx, column=col_idx + 1)
                
                color = self.get_cell_color(source_cell)
                score = self.SUBJECTIVE_SCORES.get(color, 0)
                target_cell.value = score
                target_cell.fill = copy(source_cell.fill)  # [추가] 배경색 복사
            
            # [수식 입력]
            # 컬럼 문자 가져오기
            o_col = get_column_letter(self.OBJECTIVE_SUM_COL + 1)
            e_col = get_column_letter(self.OBJECTIVE_COLS[0] + 1)
            n_col = get_column_letter(self.OBJECTIVE_COLS[-1] + 1)
            
            ae_col = get_column_letter(self.SUBJECTIVE_SUM_COL + 1)
            p_col = get_column_letter(self.SUBJECTIVE_COLS[0] + 1)
            ad_col = get_column_letter(self.SUBJECTIVE_COLS[-1] + 1)
            
            af_col = get_column_letter(self.TOTAL_SUM_COL + 1)
            
            # 타겟 시트(채점결과)에 합계 수식 입력
            target_sheet.cell(row=row_idx, column=self.OBJECTIVE_SUM_COL + 1).value = f"=SUM({e_col}{row_idx}:{n_col}{row_idx})"
            target_sheet.cell(row=row_idx, column=self.SUBJECTIVE_SUM_COL + 1).value = f"=SUM({p_col}{row_idx}:{ad_col}{row_idx})"
            target_sheet.cell(row=row_idx, column=self.TOTAL_SUM_COL + 1).value = f"={o_col}{row_idx}+{ae_col}{row_idx}"

            # 원본 시트에 참조 수식 입력
            source_sheet.cell(row=row_idx, column=self.OBJECTIVE_SUM_COL + 1).value = f"='{target_sheet_name}'!{o_col}{row_idx}"
            source_sheet.cell(row=row_idx, column=self.SUBJECTIVE_SUM_COL + 1).value = f"='{target_sheet_name}'!{ae_col}{row_idx}"
            source_sheet.cell(row=row_idx, column=self.TOTAL_SUM_COL + 1).value = f"='{target_sheet_name}'!{af_col}{row_idx}"

        # [열 너비 설정] 채점결과 시트 E열부터 (5번째 열)
        for c in range(5, max_c + 1):
             target_sheet.column_dimensions[get_column_letter(c)].width = 6

        new_wb.save(output)
        output.seek(0)
        return output
